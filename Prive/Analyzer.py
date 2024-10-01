import openpyxl
from Spotify import get_track_features


def main():
    print("***THIS IS A PERSONAL SPOTIFY ANALYZER***\n")
    file_name = "Prive/Stemlijst.xlsx"

    nested_dict = open_file(file_name)
    find_highest(nested_dict)

    
def open_file(file_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_name)

    # Get all sheet names
    sheet_names = workbook.sheetnames

    # Initialize an empty dictionary to accumulate data from all sheets
    nested_dict = {}

    # Loop through all sheets
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        # Iterate through the rows in each sheet (skipping the header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Ensure there's data in the row
                nested_dict = seperate_lines(row, nested_dict)

    return nested_dict


def seperate_lines(row, nested_dict):
    row = list(row)
    add_to_dict(nested_dict, row)
    return nested_dict


def add_to_dict(nested_dict, row):
    # Split the names by comma and trim whitespace
    names = [name.strip() for name in row[2].split(",")]  # Assuming names are in the 3rd column
    title = row[0]  # Assuming the title is in the 1st column
    artist = row[1]  # Assuming the artist is in the 2nd column
    rating = row[3]

    # For each name in the list, add the title and artist
    for name in names:
        # If the name already exists, append the title and artist to the list
        if name in nested_dict:
            nested_dict[name]["Title"].append(title)
            nested_dict[name]["Artist"].append(artist)
            nested_dict[name]["Rating"].append(rating)
        else:
            # Otherwise, create a new entry with lists for title and artist
            nested_dict[name] = {
                "Title": [title],
                "Artist": [artist],
                "Rating": [rating]
            }


def find_highest(nested_dict):
    file_name = "Prive/Analysed.xlsx"
    # Step 1: Create a list of tuples with (rating, title, artist) for all entries
    rating_list = []
    for details in nested_dict.values():
        titles = details["Title"]
        artists = details["Artist"]
        ratings = details["Rating"]

        for i in range(len(ratings)):
            # Replace '&apos;' with "'" in titles and artists
            title = titles[i].replace("&apos;", "'")
            artist = artists[i].replace("&apos;", "'")
            rating_list.append((ratings[i], title, artist))

    # Step 2: Sort the list by rating in descending order
    rating_list.sort(reverse=True, key=lambda x: x[0])

    # Step 3: Remove duplicates based on (title, artist)
    seen = set()
    unique_ratings = []
    for rating, title, artist in rating_list:
        if (title, artist) not in seen:
            seen.add((title, artist))
            unique_ratings.append((rating, title, artist))

    # Step 4: Get the top 10 unique ratings and print them
    print("Top 20 Highest Rated Songs:")
    for i in range(min(3, len(unique_ratings))):  # Ensure we don't exceed the number of available entries
        rating, title, artist = unique_ratings[i]
        print(f"{i+1}. {title} by {artist} - Rating: {rating}")
        info = get_track_features(title, artist)
        append_to_excel(file_name, info)
        print()


def append_to_excel(file_name, data):               # noqa: ARG001
    print()


main()
