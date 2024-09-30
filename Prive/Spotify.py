import openpyxl


def main():
    print("***THIS IS A PERSONAL SPOTIFY ANALYZER***\n")
    file_name = "Prive/Stemlijst.xlsx"

    nested_dict = open_file(file_name)
    print(nested_dict["Xander"], "\n")

    for key , _item in nested_dict.items():
        print(key)


def open_file(file_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_name)

    # Get all sheet names
    sheet_names = workbook.sheetnames

    # Initialize an empty dictionary to accumulate data from all sheets
    nested_dict = {}

    # Loop through all sheets
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        # Iterate through the rows in each sheet (skipping the header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Ensure there's data in the row
                nested_dict = seperate_lines(row, nested_dict)

    return nested_dict


def seperate_lines(row, nested_dict):
    row = list(row)
    add_to_dict(nested_dict, row)
    return nested_dict


def add_to_dict(nested_dict, row):
    # Split the names by comma and trim whitespace
    names = [name.strip() for name in row[2].split(",")]  # Assuming names are in the 3rd column
    title = row[0]  # Assuming the title is in the 1st column
    artist = row[1]  # Assuming the artist is in the 2nd column

    # For each name in the list, add the title and artist
    for name in names:
        # If the name already exists, append the title and artist to the list
        if name in nested_dict:
            nested_dict[name]["Title"].append(title)
            nested_dict[name]["Artist"].append(artist)
        else:
            # Otherwise, create a new entry with lists for title and artist
            nested_dict[name] = {
                "Title": [title],
                "Artist": [artist]
            }


main()
