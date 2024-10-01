import os
import openpyxl
from Spotify import get_track_features


def main():
    print("***THIS IS A PERSONAL SPOTIFY ANALYZER***\n")
    
    file_name = "Prive/Spotistats/Stemlijst.xlsx"
    nested_dict = open_file(file_name)

    output_file_name = 'song_data_analyzed.xlsx'
    initialize_excel(output_file_name)

    find_highest(nested_dict, output_file_name)


def open_file(file_name):
    """Loads the Excel file and collects data into a nested dictionary."""
    workbook = openpyxl.load_workbook(file_name)
    nested_dict = {}

    for sheet in workbook.sheetnames:
        sheet_data = workbook[sheet]
        for row in sheet_data.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Ensure there's data in the row
                add_to_dict(nested_dict, row)

    return nested_dict


def add_to_dict(nested_dict, row):
    """Adds data from a row to the nested dictionary."""
    names = [name.strip() for name in row[2].split(",")]
    title, artist, rating = row[0], row[1], row[3]

    for name in names:
        if name not in nested_dict:
            nested_dict[name] = {"Title": [], "Artist": [], "Rating": []}

        nested_dict[name]["Title"].append(title)
        nested_dict[name]["Artist"].append(artist)
        nested_dict[name]["Rating"].append(rating)


def find_highest(nested_dict, file_name):
    """Finds and saves the top 20 highest-rated songs to an Excel file."""
    rating_list = [
        (rating, title.replace("&apos;", "'"), artist.replace("&apos;", "'"))
        for details in nested_dict.values()
        for title, artist, rating in zip(details["Title"], details["Artist"], details["Rating"], strict=False)
    ]

    unique_ratings = {(title, artist): max(rating for rating, t, a in rating_list if t == title and a == artist)
                       for _rating, title, artist in rating_list}

    sorted_ratings = sorted(unique_ratings.items(), key=lambda item: item[1], reverse=True)

    count = int(input("How many songs do you want to save: "))
    print(f"Top {count} Highest Rated Songs:")
    for i, ((title, artist), rating) in enumerate(sorted_ratings[:count]):
        print(f"{i+1}. {title} by {artist} - Rating: {rating}")
        song_info = get_track_features(title, artist) + (rating,)  # type: ignore
        write_song_data_to_excel(song_info, file_name)
        print()


def initialize_excel(file_name):
    """Creates a new Excel file with headers if it doesn't exist."""
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['Title', 'Artist', 'Popularity', 'Genres', 'Danceability',
                   'Energy', 'Loudness', 'Tempo', 'Acousticness', 'Rating']
        sheet.append(headers)  # Write headers to the first row
        workbook.save(file_name)
        print(f"New file {file_name} created with headers.")
    else:
        print(f"File {file_name} already exists, appending new data.")


def write_song_data_to_excel(song, file_name):
    """Writes a single song entry to the existing Excel file."""
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    sheet.append(song)  # Append the song data to the next available row
    workbook.save(file_name)
    print(f"Song data added to {file_name} successfully!")


if __name__ == "__main__":
    main()
